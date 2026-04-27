import os
import time
import threading
from datetime import datetime, timedelta
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from src.pitherm.smtp_client import SMTPClient
from email.mime.application import MIMEApplication
import csv
import shutil

_last_report_week = None
_excel_lock = threading.Lock()
BASE_LOG_DIR = "logs"
CURRENT_DIR = os.path.join(BASE_LOG_DIR, "current")
ARCHIVE_DIR = os.path.join(BASE_LOG_DIR, "archive")

def get_week_str(dt):
    year, week, _ = dt.isocalendar()
    return f"{year}_W{week:02d}"

def get_active_sheet(workbook):
    worksheet = workbook.active

    if worksheet is None:
        worksheet = workbook.create_sheet()

    return worksheet

def find_weekly_report_file(week_str):
    filename = f"temp_log_{week_str}.xlsx"

    for directory in (CURRENT_DIR, ARCHIVE_DIR):
        path = os.path.join(directory, filename)
        if os.path.exists(path):
            return path

    return None

def log_to_csv_fallback(temp, hum):
    ensure_log_directories()

    week_str = get_week_str(datetime.now())

    fallback_file = os.path.join(CURRENT_DIR, f"fallback_{week_str}.csv")
    now = datetime.now()
    file_exists = os.path.exists(fallback_file)

    with open(fallback_file, mode="a", newline="") as file:
        writer = csv.writer(file)

        if not file_exists:
            writer.writerow([
                "Date", 
                "Time", 
                "Temperature (°C)", 
                "Humidity (%)"
            ])

        writer.writerow([
            now.strftime("%Y-%m-%d"),
            now.strftime("%H:%M:%S"),
            temp,
            hum
        ])
    print("[FALLBACK] Logged reading to CSV.")

def ensure_log_directories():
    os.makedirs(CURRENT_DIR, exist_ok=True)
    os.makedirs(ARCHIVE_DIR, exist_ok=True)

def archive_old_logs():
    ensure_log_directories()

    current_week_str = get_week_str(datetime.now())

    for file in os.listdir(CURRENT_DIR):
        if (
            (file.startswith("temp_log_") and file.endswith(".xlsx")) or
            (file.startswith("fallback_") and file.endswith(".csv"))
        ):
            if file.startswith("temp_log_"):
                file_week = file.replace("temp_log_", "").replace(".xlsx", "")
            else:
                file_week = file.replace("fallback_", "").replace(".csv", "")

            if file_week != current_week_str:
                src_path = os.path.join(CURRENT_DIR, file)
                dst_path = os.path.join(ARCHIVE_DIR, file)

                shutil.move(src_path, dst_path)
                print(f"[ARCHIVE] Moved {file} to archive.")

def log_to_excel(temp, hum):
    ensure_log_directories()

    with _excel_lock:
        archive_old_logs()

        week_str = get_week_str(datetime.now())
        filename = os.path.join(CURRENT_DIR, f"temp_log_{week_str}.xlsx")

        try:
            try:
                wb = load_workbook(filename)
                ws = get_active_sheet(wb)
            except FileNotFoundError:
                wb = Workbook()
                ws = get_active_sheet(wb)
                ws.title = "Weekly Readings"
                ws.append(["Date", "Time", "Temperature (°C)", "Humidity (%)"])
                for col in range(1, 5):
                    ws[f"{get_column_letter(col)}1"].font = Font(bold=True)
            now = datetime.now()
            ws.append([
                now.strftime("%Y-%m-%d"), 
                now.strftime("%H:%M:%S"), 
                temp, 
                hum
            ])
            wb.save(filename)
        except Exception as e:
            print("[CRITICAL] Excel logging failed. Switching to CSV Fallback:", e)
            log_to_csv_fallback(temp, hum)

def send_weekly_report(now=None, sender=None):
    if now is None:
        now = datetime.now()

    with _excel_lock:
        report_date = now - timedelta(days=7)
        week_str = get_week_str(report_date)
        filename = find_weekly_report_file(week_str)

    if filename is None:
        print(f"[WARN] No Excel file to send for {week_str}.")
        return False

    subject = f"Weekly Temp Report - {week_str}"
    body = f"""
    <p>Attached is the temperature and humidity log for {week_str}.</p>
    <p>- Raspberry Pi Monitor</p>
    """

    with open(filename, 'rb') as f:
        attachment = MIMEApplication(
            f.read(),
            _subtype='vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        attachment.add_header(
            'Content-Disposition',
            'attachment',
            filename=os.path.basename(filename)
        )

    if sender is None:
        sender = SMTPClient()
    
    return sender.send(
        subject, 
        body, 
        is_html=True, 
        attachment=attachment
    )

def check_and_send_weekly_report(now=None, sender=None):
    global _last_report_week

    if now is None:
        now = datetime.now()

    current_week = now.isocalendar()[:2]

    if now.weekday() == 0 and now.hour >= 7 and _last_report_week != current_week:
        if send_weekly_report(now=now, sender=sender):
            _last_report_week = current_week

def run_scheduler():
    while True:
        check_and_send_weekly_report()
        time.sleep(60)

def start_scheduler():
    thread = threading.Thread(target=run_scheduler, daemon=True)
    thread.start()
